#!/usr/bin/env python
"""
Convert monthly open-end fund reports (Thong tu 98/2020/TT-BTC, xlsx) to tidy CSVs.

Usage:
  python scripts/fund_report/fund_reports_update.py                    # all funds' raw/ dirs
  python scripts/fund_report/fund_reports_update.py <file.xlsx> ...    # specific files
  python scripts/fund_report/fund_reports_update.py --check            # compare without writing

File names must follow the standard <FUND>_<YYYY>_<MM>.xlsx pattern,
e.g. DCDS_2026_07.xlsx. The fund id is taken from the file name and each
fund is written to its own directory:

  public/data/<FUND>/raw/<FUND>_<YYYY>_<MM>.xlsx
  public/data/<FUND>/tidied/:
    tidy_assets.csv      <- BCTaiSan           (balance sheet)
    tidy_income.csv      <- BCKetQuaHoatDong   (profit & loss)
    tidy_portfolio.csv   <- BCDanhMucDauTu     (investment portfolio)
    tidy_indicators.csv  <- Khac               (other indicators)
    tidy_borrowing.csv   <- BCHoatDongVay      (borrowing / repo)
    tidy_metadata.json   <- TONGQUAN headers   (per-file metadata)
    tidied_index.json                          (audit manifest)

Rules:
  - line_item is the English part of the "Tieng Viet\\nEnglish" cell.
    Cells without an English line keep their original text.
  - Drop rows whose content column is "...", "…" or empty, and rows that
    have neither a code nor any numeric value.
  - BCTaiSan column F ("%/cung ky nam truoc") is dropped on purpose:
    values are stale leftovers that do not match D/E (checked 08/2026).
  - Duplicates on (key, see KEYS below) keep the row with the newest asOf,
    so re-running a file or processing overlapping months never duplicates.
  - Every source file is validated for internal consistency before any
    output is written (structure guard + number reconciliations).

Depends on: openpyxl, pandas (both already installed on the dev machine).
"""

import argparse
import calendar
import json
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent.parent
DATA_DIR = ROOT / "public" / "data"

# Standard report file name: <FUND>_<YYYY>_<MM>.xlsx
FUND_FILE_RE = re.compile(r"^([A-Za-z0-9]+)_(\d{4})_(\d{2})\.xlsx$")

TABLES = ["assets", "income", "portfolio", "indicators", "borrowing"]

COLS = {
    "assets": ["code", "line_item", "period_end", "value", "asOf"],
    "income": ["code", "line_item", "period_end", "measure", "value", "asOf"],
    "portfolio": [
        "period_end", "section", "code", "ticker", "quantity",
        "market_price", "value", "weight", "asOf",
    ],
    "indicators": ["code", "line_item", "period_end", "measure", "value", "asOf"],
    "borrowing": ["period_end", "item", "amount", "balance_nav_pct", "asOf"],
}

KEYS = {
    "assets": ["code", "line_item", "period_end"],
    "income": ["code", "line_item", "period_end", "measure"],
    "portfolio": ["period_end", "code"],
    "indicators": ["code", "line_item", "period_end", "measure"],
    "borrowing": ["period_end", "item"],
}

SHEET_NAMES = {
    "assets": {"prefix": "BCTaiSan"},
    "income": {"prefix": "BCKetQuaHoatDong"},
    "portfolio": {"prefix": "BCDanhMucDauTu"},
    "indicators": {"prefix": "Khac"},
    "borrowing": {"prefix": "BCHoatDongVay"},
}

# Value columns (1-based) per table, used to decide whether a row is kept
# and where to read numbers from.
VALUE_COLS = {
    "assets": (4, 5),
    "income": (4, 5, 6),
    "portfolio": (4, 5, 6, 7),
    "indicators": (4, 5),
    "borrowing": (6, 8, 10),
}


def fund_from_name(name):
    """Fund id from the standard '<FUND>_<YYYY>_<MM>.xlsx' file name."""
    m = FUND_FILE_RE.match(name)
    if not m:
        raise ValueError(
            f"file '{name}' does not match the standard name "
            f"<FUND>_<YYYY>_<MM>.xlsx (e.g. DCDS_2026_07.xlsx)"
        )
    fund = m.group(1).upper()
    year, month = int(m.group(2)), int(m.group(3))
    if not (2000 <= year <= 2100 and 1 <= month <= 12):
        raise ValueError(f"file '{name}' has an invalid year/month ({year}-{month:02d})")
    return fund


def fund_dirs(fund):
    base = DATA_DIR / fund
    return base / "raw", base / "tidied"


def all_raw_files():
    return sorted(DATA_DIR.glob("*/raw/*.xlsx"))


def cell_str(v):
    if v is None:
        return ""
    return str(v)


def last_day_of_month(year, month):
    return date(year, month, calendar.monthrange(year, month)[1])


def parse_vn_date(text):
    """Parse 'Ngay 30 thang 06 nam 2026' or 'Thang 06 nam 2026' (last day)."""
    if isinstance(text, datetime):
        return text.date()
    if isinstance(text, date):
        return text
    if not isinstance(text, str):
        return None
    # Old reports (2018-2020) store Vietnamese in NFD (decomposed) form:
    # 'a' + U+0300 instead of 'a'. Normalize so the regex matches both.
    text = unicodedata.normalize("NFC", text)
    m = re.search(r"Ng\u00e0y\s+(\d{1,2})\s+th\u00e1ng\s+(\d{1,2})\s+n\u0103m\s+(\d{4})", text, re.IGNORECASE)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    m = re.search(r"Th\u00e1ng\s+(\d{1,2})\s+n\u0103m\s+(\d{4})", text, re.IGNORECASE)
    if m:
        return last_day_of_month(int(m.group(2)), int(m.group(1)))
    return None


def _ascii_view(s):
    """s with typographic punctuation normalized, for isascii() checks only.

    The report template types 'Fund's' with a curly apostrophe (U+2019) on
    the English lines. The original text is kept in the output; this is only
    used to decide where the English part starts.
    """
    return (
        s.replace("\u2019", "'").replace("\u2018", "'")
        .replace("\u201c", '"').replace("\u201d", '"')
        .replace("\u2013", "-").replace("\u2014", "-")
    )


def english_part(text):
    """English part of a 'Tieng Viet\\nEnglish' cell, else the original text.

    Multi-line cells carry "VN\\nEN" (template convention): take everything
    after the first line. A few labels are merged on one line, e.g.
    "TAI SAN ASSETS": scan for the first token whose whole tail is pure
    ASCII. Vietnamese words written without diacritics ('quay', 'danh'...)
    are skipped because their tail still contains diacritic words, while a
    formula suffix like '(=I+II)' is rejected because a single-token tail
    must contain letters.
    """
    if not isinstance(text, str):
        return ""
    text = text.strip()
    parts = text.split("\n")
    if len(parts) > 1:
        # English starts at the first line that is pure ASCII. Vietnamese
        # lines sometimes wrap over two lines (e.g. the "Khac" sheet), so
        # "take everything after line 1" is not enough.
        for i, line in enumerate(parts):
            if line.strip() and _ascii_view(line).isascii():
                return " ".join(p.strip() for p in parts[i:] if p.strip())
        return text
    tokens = text.split()
    for i, tok in enumerate(tokens):
        if not _ascii_view(tok).isascii():
            continue
        tail = tokens[i:]
        if not all(_ascii_view(t).isascii() for t in tail):
            continue
        if len(tail) >= 2:
            en = re.sub(r"^[^A-Za-z0-9]+", "", " ".join(tail))
            # formulas like '(=III + IV)' are all-caps with no lowercase letter
            if en and re.search(r"[a-z]", en):
                return en
            return text
        core = re.sub(r"^[^A-Za-z0-9]+|[^A-Za-z0-9]+$", "", tail[0])
        if core.isalnum():
            return tail[0]
    return text


def is_spacer(text):
    if not isinstance(text, str):
        return False
    t = text.strip()
    return t in ("...", "\u2026") or t.startswith("...")


def find_sheet(wb, spec):
    for ws in wb.worksheets:
        if "exact" in spec and ws.title == spec["exact"]:
            return ws
        if "prefix" in spec and ws.title.startswith(spec["prefix"]):
            return ws
    return None


def find_header_row(ws):
    """Row whose column A starts with 'STT' (the data header row)."""
    for row in ws.iter_rows(min_row=1, max_row=40):
        a = row[0].value
        if isinstance(a, str) and a.strip().startswith("STT"):
            return row[0].row
    raise ValueError("header row (STT) not found")


def meta_value(ws, label):
    """Label in column A, value in column C (metadata block of every sheet)."""
    for r in range(1, 20):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().startswith(label):
            return ws.cell(r, 3).value
    return None


def numeric_or_none(v):
    return None if v is None else float(v)


def table_rows(ws, table, header, as_of, keep_text_rows=False):
    vcols = VALUE_COLS[table]
    rows = []
    for r in range(header + 1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if not isinstance(b, str) or not b.strip():
            continue
        if is_spacer(b):
            continue
        values = [ws.cell(r, c).value for c in vcols]
        code = cell_str(ws.cell(r, 3).value)
        if not code and all(v is None for v in values) and not keep_text_rows:
            continue
        rows.append((r, b, code, values))
    return rows


def _require_date(value, what, sheet_title):
    if value is None:
        raise ValueError(f"cannot parse {what} date in sheet '{sheet_title}'")
    return value


def build_assets(ws, header, as_of):
    cur = _require_date(parse_vn_date(ws.cell(header, 4).value), "current period", ws.title)
    prev = _require_date(parse_vn_date(ws.cell(header, 5).value), "previous period", ws.title)
    out = []
    for _r, b, code, (d, e) in table_rows(ws, "assets", header, as_of):
        label = english_part(b)
        # The 2018-2020 template numbers the sections differently, so the
        # NAV row reads 'Net Asset Value ( = I.8 - II.3)' there while 2021+
        # uses '= I.10 - II.4'. Normalize so the same line item merges.
        label = label.replace("= I.8 - II.3", "= I.10 - II.4")
        out.append([code, label, cur.isoformat(), cell_str(d), as_of])
        out.append([code, label, prev.isoformat(), cell_str(e), as_of])
    return out


def build_income(ws, header, as_of):
    cur = _require_date(parse_vn_date(ws.cell(header, 4).value), "current period", ws.title)
    prev = _require_date(parse_vn_date(ws.cell(header, 5).value), "previous period", ws.title)
    out = []
    for _r, b, code, (d, e, f) in table_rows(ws, "income", header, as_of):
        label = english_part(b)
        out.append([code, label, cur.isoformat(), "month", cell_str(d), as_of])
        out.append([code, label, prev.isoformat(), "month", cell_str(e), as_of])
        out.append([code, label, cur.isoformat(), "ytd", cell_str(f), as_of])
    return out


def build_indicators(ws, header, as_of):
    cur = _require_date(parse_vn_date(ws.cell(header, 4).value), "current period", ws.title)
    prev = _require_date(parse_vn_date(ws.cell(header, 5).value), "previous period", ws.title)
    out = []
    for _r, b, code, (d, e) in table_rows(ws, "indicators", header, as_of):
        label = english_part(b)
        out.append([code, label, cur.isoformat(), "month", cell_str(d), as_of])
        out.append([code, label, prev.isoformat(), "month", cell_str(e), as_of])
    return out


def build_portfolio(ws, header, as_of):
    period = _require_date(parse_vn_date(ws.cell(5, 1).value), "reporting", ws.title)
    section = ""
    out = []
    for _r, b, code, (d, e, f, g) in table_rows(ws, "portfolio", header, as_of, keep_text_rows=True):
        a_str = cell_str(ws.cell(_r, 1).value).strip()
        # Section headers have a roman-numeral STT. In 2021-2022 reports the
        # code column (C) is empty on every row, so the header test must not
        # require a code.
        if re.fullmatch(r"[IVX]+", a_str) and b.strip():
            section = english_part(b)
            ticker = ""
        elif b.strip().startswith("T\u1ed4NG"):
            ticker = ""
        else:
            ticker = english_part(b)
        out.append([
            period.isoformat(), section, code, ticker,
            cell_str(d), cell_str(e), cell_str(f), cell_str(g), as_of,
        ])
        # The 2018-2020 reports append a second table (futures positions) to
        # this sheet below the grand total. Stop at "Total value of portfolio"
        # so that table never leaks into the portfolio data.
        if "Total value of portfolio" in b:
            break
    return out


def build_borrowing(ws, header, as_of):
    period = _require_date(parse_vn_date(ws.cell(5, 1).value), "reporting", ws.title)
    out = []
    for _r, b, _code, (f, _h, j) in table_rows(ws, "borrowing", header, as_of, keep_text_rows=True):
        out.append([period.isoformat(), english_part(b), cell_str(f), cell_str(j), as_of])
    return out


BUILDERS = {
    "assets": build_assets,
    "income": build_income,
    "portfolio": build_portfolio,
    "indicators": build_indicators,
    "borrowing": build_borrowing,
}


# --------------------------------------------------------------------------
# Validation: internal consistency of a source file.
# --------------------------------------------------------------------------

def _value_map(ws, header, col):
    """code -> numeric value of `col` for the data rows of a sheet."""
    out = {}
    for r in range(header + 1, ws.max_row + 1):
        code = ws.cell(r, 3).value
        if code is None:
            continue
        raw = ws.cell(r, col).value
        if is_spacer(raw):
            continue
        # Some old reports put a text label in a numeric column (e.g. the
        # futures-position table inside the portfolio sheet). Never crash on
        # that; only numeric cells take part in the reconciliation.
        if isinstance(raw, (int, float)):
            out[str(code).strip()] = float(raw)
        elif isinstance(raw, str):
            try:
                out[str(code).strip()] = float(raw)
            except ValueError:
                pass
    return out


def _check_equal(name, expected, actual, tol=0.01):
    if expected is None or actual is None:
        print(f"  [skip] {name}: missing value")
        return
    if abs(expected - actual) > tol:
        raise ValueError(f"{name}: {expected} != {actual}")


def validate_file(wb, sheets):
    """Raise ValueError on any inconsistency; print checks passed."""
    print("  validation:")

    assets_ws = sheets.get("assets")
    if assets_ws is not None:
        h = find_header_row(assets_ws)
        cur = _value_map(assets_ws, h, 4)   # current period
        prev = _value_map(assets_ws, h, 5)  # previous period
        _check_equal("assets: total assets = liabilities + NAV (current)",
                     cur.get("2212"),
                     None if (cur.get("2216") is None or cur.get("2217") is None)
                     else cur["2216"] + cur["2217"])
        _check_equal("assets: total assets = liabilities + NAV (previous)",
                     prev.get("2212"),
                     None if (prev.get("2216") is None or prev.get("2217") is None)
                     else prev["2216"] + prev["2217"])

    income_ws = sheets.get("income")
    if income_ws is not None:
        h = find_header_row(income_ws)
        m = _value_map(income_ws, h, 4)  # current month
        _check_equal("income: net income = income - expenses (month)",
                     m.get("2233"),
                     None if (m.get("2220") is None or m.get("2224") is None)
                     else m["2220"] - m["2224"])
        _check_equal("income: change due to investment = net + gain/loss (month)",
                     m.get("2237"),
                     None if (m.get("2233") is None or m.get("2234") is None)
                     else m["2233"] + m["2234"])
        # NAV continuity across sheets: P&L beginning-of-period == balance sheet end of previous period
        if assets_ws is not None:
            ha = find_header_row(assets_ws)
            prev_nav = _value_map(assets_ws, ha, 5).get("2217")
            _check_equal("cross-sheet: P&L NAV beginning == balance sheet NAV previous period",
                         m.get("2238"), prev_nav)

    port_ws = sheets.get("portfolio")
    if port_ws is not None:
        h = find_header_row(port_ws)
        pm = _value_map(port_ws, h, 6)  # column F = value (VND)
        total_2247 = pm.get("2247")
        child_sum = sum(v for k, v in pm.items() if k.startswith("2246."))
        _check_equal("portfolio: listed shares total == sum of stocks", total_2247, child_sum)
        _check_equal("portfolio: grand total == sections total",
                     pm.get("2263"),
                     None if pm.get("2255") is None or pm.get("2257") is None or pm.get("2262") is None
                     else pm["2255"] + pm["2257"] + pm["2262"])
    print("  [ok] all checks passed")


# --------------------------------------------------------------------------
# Merge / dedupe / write.
# --------------------------------------------------------------------------

def natural_parts(v):
    if v is None or v == "":
        return ()
    return tuple(int(p) if p.isdigit() else p for p in re.split(r"\.", str(v)))


def backfill_codes(df, table):
    """Fill empty codes from an unambiguous {line_item: code} map.

    The 2021-2022 reports left the code column empty in BCTaiSan. Rows whose
    English line_item maps to exactly one code in the coded era (2023+) are
    given that code. Ambiguous labels (e.g. 'Other payables' = 2215.17 and
    2215.17.4) stay empty rather than risk a wrong assignment.
    """
    if table not in ("assets", "income", "indicators"):
        return df
    coded = df[df["code"] != ""]
    if coded.empty:
        return df
    label_to_codes = {}
    for li, code in zip(coded["line_item"], coded["code"]):
        label_to_codes.setdefault(li, set()).add(code)
    unambiguous = {
        li: next(iter(codes))
        for li, codes in label_to_codes.items()
        if len(codes) == 1
    }
    if not unambiguous:
        return df
    empty = df["code"] == ""
    df.loc[empty, "code"] = df.loc[empty, "line_item"].map(unambiguous).fillna("")
    return df


def merge_table(table, new_rows, tidy_dir):
    """Merge new rows into the table's CSV content (or a fresh frame)."""
    path = tidy_dir / f"tidy_{table}.csv"
    cols = COLS[table]
    new_df = pd.DataFrame(new_rows, columns=cols)
    if path.exists():
        old_df = pd.read_csv(path, dtype=str, keep_default_na=False)
        df = pd.concat([old_df, new_df], ignore_index=True)
    else:
        df = new_df
    df = backfill_codes(df, table)
    if table == "portfolio":
        # Portfolio rows carry no stable natural key: in the 2021-2022 era the
        # code column is empty on every row (section headers and totals
        # included), so keying on (period_end, code) collapses them. Dedupe
        # exact duplicates only; re-running a file is idempotent, distinct
        # rows always survive.
        df = df.drop_duplicates(keep="last")
        df = df.sort_values(["period_end"], kind="stable")
        return df
    keys = KEYS[table]
    if table in ("assets", "income", "indicators"):
        # In the 2021-2022 era the code column is empty on many rows, and two
        # genuinely different line items can share the same English label
        # (e.g. 'Other payables' parent vs sub-item). Keying such rows by
        # (code, line_item, period) would collapse them and keep a wrong value.
        # For empty-code rows key on (line_item, period, value): identical
        # rows from overlapping months merge (newest asOf), distinct rows
        # survive.
        coded = df[df["code"] != ""]
        empty = df[df["code"] == ""]
        if not coded.empty:
            coded = coded.sort_values(keys + ["asOf"], kind="stable") \
                         .drop_duplicates(subset=keys, keep="last")
        if not empty.empty:
            ekeys = [k for k in keys if k != "code"]
            empty = empty.sort_values(ekeys + ["value", "asOf"], kind="stable") \
                         .drop_duplicates(subset=ekeys + ["value"], keep="last")
        df = pd.concat([coded, empty], ignore_index=True)
    else:
        df = df.sort_values(keys + ["asOf"], kind="stable")
        df = df.drop_duplicates(subset=keys, keep="last")
    if table == "borrowing":
        df = df.sort_values(["period_end", "item"], kind="stable")
    else:
        df["_nk"] = df["code"].map(natural_parts)
        df = df.sort_values(["period_end", "_nk"], kind="stable")
        df = df.drop(columns=["_nk"])
    return df


def upsert_table(table, new_rows, tidy_dir):
    df = merge_table(table, new_rows, tidy_dir)
    path = tidy_dir / f"tidy_{table}.csv"
    df.to_csv(path, index=False, encoding="utf-8", lineterminator="\n")
    return len(df)


def build_manifests(meta_per_file, latest_as_of, latest_period, fund):
    index = {
        "fund": fund,
        "sheets": TABLES,
        "source_files": sorted(meta_per_file.keys()),
        "latest_period": latest_period,
        "latest_as_of": latest_as_of,
        "updated": date.today().isoformat(),
    }
    metadata = {
        "fund_id": fund,
        "files": [meta_per_file[name] for name in sorted(meta_per_file)],
    }
    return index, metadata


def write_manifests(meta_per_file, tidy_dir, fund):
    latest_as_of = None
    latest_period = None
    for tbl in TABLES:
        path = tidy_dir / f"tidy_{tbl}.csv"
        if not path.exists():
            continue
        df = pd.read_csv(path, dtype=str)
        for v in df.get("asOf", []):
            if v and (latest_as_of is None or v > latest_as_of):
                latest_as_of = v
        for v in df.get("period_end", []):
            if v and (latest_period is None or v > latest_period):
                latest_period = v

    index, metadata = build_manifests(meta_per_file, latest_as_of, latest_period, fund)
    with open(tidy_dir / "tidied_index.json", "w", encoding="utf-8") as fh:
        json.dump(index, fh, ensure_ascii=False, indent=2)
        fh.write("\n")
    with open(tidy_dir / "tidy_metadata.json", "w", encoding="utf-8") as fh:
        json.dump(metadata, fh, ensure_ascii=False, indent=2)
        fh.write("\n")


def process_file(path):
    fund = fund_from_name(path.name)
    raw_dir, _ = fund_dirs(fund)
    if path.parent.resolve() != raw_dir.resolve():
        raise ValueError(
            f"file '{path.name}' names fund '{fund}' but is not in {raw_dir} — "
            f"a misnamed or misplaced file would be written into the wrong fund's tidied files"
        )
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"== {path.name} (fund {fund})")
    sheets = {}
    for table in TABLES:
        ws = find_sheet(wb, SHEET_NAMES[table])
        if ws is None:
            print(f"  [warn] sheet for {table} not found, skipped")
            continue
        sheets[table] = ws
    if not sheets:
        raise ValueError(f"no known sheets in {path.name}")

    validate_file(wb, sheets)

    meta_ws = sheets.get("assets") or sheets.get("income")
    as_of_raw = meta_value(meta_ws, "Ng\u00e0y l\u1eadp b\u00e1o c\u00e1o")
    as_of = parse_vn_date(as_of_raw)
    if as_of is None:
        raise ValueError(f"cannot parse asOf from '{as_of_raw}' in {path.name}")

    metadata = {
        "source_file": path.name,
        "as_of": as_of.isoformat(),
        "fund_name_vi": cell_str(meta_value(meta_ws, "T\u00ean Qu\u1ef9:")).strip(),
        "fund_name_en": cell_str(meta_value(meta_ws, "Fund name:")).strip(),
        "company_vi": cell_str(meta_value(meta_ws, "T\u00ean C\u00f4ng ty qu\u1ea3n l\u00fd qu\u1ef9:")).strip(),
        "company_en": cell_str(meta_value(meta_ws, "Fund Management Company:")).strip(),
        "bank_vi": cell_str(meta_value(meta_ws, "T\u00ean ng\u00e2n h\u00e0ng gi\u00e1m s\u00e1t:")).strip(),
        "bank_en": cell_str(meta_value(meta_ws, "Supervising bank:")).strip(),
    }

    rows_by_table = {}
    for table, ws in sheets.items():
        header = find_header_row(ws)
        rows = BUILDERS[table](ws, header, as_of.isoformat())
        rows_by_table[table] = rows
        print(f"  {table}: {len(rows)} rows from {path.name}")

    wb.close()
    return rows_by_table, metadata


def run_check():
    """Rebuild every table as a fresh run would and compare with disk."""
    files = all_raw_files()
    if not files:
        print("no xlsx files found in public/data/*/raw/", file=sys.stderr)
        return True
    any_diff = False
    for fund, fund_files in group_by_fund(files).items():
        tidy_dir = fund_dirs(fund)[1]
        print(f"== check fund {fund}")
        all_rows = {t: [] for t in TABLES}
        meta_per_file = {}
        for f in fund_files:
            rows_by_table, metadata = process_file(f)
            meta_per_file[f.name] = metadata
            for t, rows in rows_by_table.items():
                all_rows[t].extend(rows)
        for table in TABLES:
            path = tidy_dir / f"tidy_{table}.csv"
            if not path.exists():
                print(f"{table}: MISSING on disk")
                any_diff = True
                continue
            expected = merge_table(table, all_rows[table], tidy_dir)
            disk = pd.read_csv(path, dtype=str, keep_default_na=False)
            exp = sorted(tuple(str(v) for v in r) for r in expected.values.tolist())
            act = sorted(tuple(str(v) for v in r) for r in disk.values.tolist())
            if exp == act:
                print(f"{table}: OK ({len(act)} rows)")
            else:
                any_diff = True
                print(f"{table}: WOULD CHANGE ({len(exp)} rebuilt vs {len(act)} on disk)")

        # manifests: rebuild from the source files and compare with disk
        # (the 'updated' field is the run date, not data, so it is excluded)
        latest_as_of = None
        latest_period = None
        for table in TABLES:
            df_path = tidy_dir / f"tidy_{table}.csv"
            if not df_path.exists():
                continue
            df = pd.read_csv(df_path, dtype=str, keep_default_na=False)
            for v in df.get("asOf", []):
                if v and (latest_as_of is None or v > latest_as_of):
                    latest_as_of = v
            for v in df.get("period_end", []):
                if v and (latest_period is None or v > latest_period):
                    latest_period = v
        exp_index, exp_metadata = build_manifests(meta_per_file, latest_as_of, latest_period, fund)
        for name, exp in (("tidied_index.json", exp_index), ("tidy_metadata.json", exp_metadata)):
            disk_path = tidy_dir / name
            if not disk_path.exists():
                print(f"{name}: MISSING on disk")
                any_diff = True
                continue
            disk = json.loads(disk_path.read_text(encoding="utf-8"))
            if name == "tidied_index.json":
                exp = {k: v for k, v in exp.items() if k != "updated"}
                disk = {k: v for k, v in disk.items() if k != "updated"}
            if exp == disk:
                print(f"{name}: OK")
            else:
                any_diff = True
                print(f"{name}: WOULD CHANGE")
    return any_diff


def group_by_fund(files):
    """files grouped by fund id derived from each file name."""
    groups = {}
    for f in files:
        fund = fund_from_name(f.name)
        groups.setdefault(fund, []).append(f)
    return groups


def main():
    ap = argparse.ArgumentParser(description="Convert monthly fund reports to tidy CSVs")
    ap.add_argument("paths", nargs="*", help="xlsx files to process (default: all funds' raw/)")
    ap.add_argument("--check", action="store_true", help="compare current tidied files, write nothing")
    args = ap.parse_args()

    if args.check:
        sys.exit(0 if not run_check() else 1)

    if args.paths:
        files = [Path(p) for p in args.paths]
    else:
        files = all_raw_files()
    if not files:
        print("no xlsx files found in public/data/*/raw/", file=sys.stderr)
        sys.exit(1)

    for fund, fund_files in sorted(group_by_fund(files).items()):
        tidy_dir = fund_dirs(fund)[1]
        tidy_dir.mkdir(parents=True, exist_ok=True)
        print(f"== fund {fund}: {len(fund_files)} file(s)")

        all_rows = {t: [] for t in TABLES}
        meta_per_file = {}
        for f in fund_files:
            rows_by_table, metadata = process_file(f)
            meta_per_file[f.name] = metadata
            for t, rows in rows_by_table.items():
                all_rows[t].extend(rows)

        print("  writing tidied files")
        for t in TABLES:
            n = upsert_table(t, all_rows[t], tidy_dir)
            print(f"    tidy_{t}.csv: {n} rows total")

        write_manifests(meta_per_file, tidy_dir, fund)
        print("  tidied_index.json + tidy_metadata.json written")


if __name__ == "__main__":
    main()
